from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
import logging
import PyPDF2
import io
import requests
from typing import Optional
import docx
import openpyxl
from bs4 import BeautifulSoup

router = APIRouter()
logger = logging.getLogger(__name__)

class ParseDocumentRequest(BaseModel):
    url: str

class ParseDocumentResponse(BaseModel):
    success: bool
    content: str
    title: str
    doc_type: str
    error: Optional[str] = None
    is_truncated: bool = False
    total_length: int = 0

def extract_pdf_content(file_bytes: bytes) -> str:
    """Extract text from PDF bytes"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting PDF: {e}")
        raise

def extract_docx_content(file_bytes: bytes) -> str:
    """Extract text from DOCX bytes"""
    try:
        doc = docx.Document(io.BytesIO(file_bytes))
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting DOCX: {e}")
        raise

def extract_xlsx_content(file_bytes: bytes) -> str:
    """Extract text from XLSX bytes"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n"
            
            # Get all rows
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting XLSX: {e}")
        raise

def extract_html_content(html_bytes: bytes) -> str:
    """Extract text from HTML bytes"""
    try:
        soup = BeautifulSoup(html_bytes, 'html.parser')
        
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Get text
        text = soup.get_text()
        
        # Break into lines and remove leading/trailing space
        lines = (line.strip() for line in text.splitlines())
        # Break multi-headlines into a line each
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        # Drop blank lines
        text = '\n'.join(chunk for chunk in chunks if chunk)
        
        return text
    except Exception as e:
        logger.error(f"Error extracting HTML: {e}")
        raise

@router.post("/document/parse", response_model=ParseDocumentResponse)
async def parse_document(request: ParseDocumentRequest):
    """
    Parse document from URL and extract text content.
    Supports: PDF, DOCX, XLSX, HTML
    """
    try:
        url = request.url
        logger.info(f"Parsing document from URL: {url}")
        
        # Determine document type from URL
        url_lower = url.lower()
        doc_type = "unknown"
        
        if url_lower.endswith('.pdf'):
            doc_type = "pdf"
        elif url_lower.endswith('.docx') or url_lower.endswith('.doc'):
            doc_type = "docx"
        elif url_lower.endswith('.xlsx') or url_lower.endswith('.xls'):
            doc_type = "xlsx"
        elif url_lower.endswith('.html') or url_lower.endswith('.htm'):
            doc_type = "html"
        else:
            # Try to detect from content-type
            try:
                head_response = requests.head(url, timeout=5, allow_redirects=True)
                content_type = head_response.headers.get('content-type', '').lower()
                
                if 'pdf' in content_type:
                    doc_type = "pdf"
                elif 'word' in content_type or 'officedocument.wordprocessing' in content_type:
                    doc_type = "docx"
                elif 'excel' in content_type or 'spreadsheet' in content_type:
                    doc_type = "xlsx"
                elif 'html' in content_type:
                    doc_type = "html"
            except:
                pass
        
        # Download the document
        logger.info(f"Downloading document (type: {doc_type})...")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        file_bytes = response.content
        
        # Extract content based on type
        content = ""
        title = url.split('/')[-1]  # Default title from filename
        
        if doc_type == "pdf":
            content = extract_pdf_content(file_bytes)
            logger.info(f"Extracted {len(content)} characters from PDF")
        elif doc_type == "docx":
            content = extract_docx_content(file_bytes)
            logger.info(f"Extracted {len(content)} characters from DOCX")
        elif doc_type == "xlsx":
            content = extract_xlsx_content(file_bytes)
            logger.info(f"Extracted {len(content)} characters from XLSX")
        elif doc_type == "html":
            content = extract_html_content(file_bytes)
            logger.info(f"Extracted {len(content)} characters from HTML")
        else:
            # Try all parsers
            logger.info("Unknown document type, trying all parsers...")
            for parser_name, parser_func in [
                ("PDF", extract_pdf_content),
                ("DOCX", extract_docx_content),
                ("XLSX", extract_xlsx_content),
                ("HTML", extract_html_content)
            ]:
                try:
                    content = parser_func(file_bytes)
                    doc_type = parser_name.lower()
                    logger.info(f"Successfully parsed as {parser_name}")
                    break
                except:
                    continue
            
            if not content:
                raise HTTPException(
                    status_code=400,
                    detail="Unable to parse document. Supported formats: PDF, DOCX, XLSX, HTML"
                )
        
        if not content.strip():
            raise HTTPException(
                status_code=400,
                detail="Document appears to be empty or content could not be extracted"
            )
        
        # Truncate content if too large (max ~15000 characters = ~3750 tokens)
        # This prevents Groq API token limit errors
        MAX_CONTENT_LENGTH = 15000
        total_length = len(content)
        is_truncated = False
        
        if total_length > MAX_CONTENT_LENGTH:
            is_truncated = True
            # Keep first 70% and last 10% to preserve context
            first_part_length = int(MAX_CONTENT_LENGTH * 0.7)
            last_part_length = int(MAX_CONTENT_LENGTH * 0.1)
            
            content = (
                content[:first_part_length] + 
                f"\n\n... [Content truncated - showing first {first_part_length} and last {last_part_length} characters of {total_length} total] ...\n\n" +
                content[-last_part_length:]
            )
            logger.info(f"Truncated content from {total_length} to {len(content)} characters")
        
        return ParseDocumentResponse(
            success=True,
            content=content,
            title=title,
            doc_type=doc_type,
            is_truncated=is_truncated,
            total_length=total_length
        )
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Error downloading document: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to download document: {str(e)}")
    except Exception as e:
        logger.error(f"Error parsing document: {e}")
        raise HTTPException(status_code=500, detail=f"Error parsing document: {str(e)}")

@router.post("/document/upload-parse")
async def parse_uploaded_document(file: UploadFile = File(...)):
    """
    Parse uploaded document file and extract text content.
    Supports: PDF, DOCX, XLSX
    """
    try:
        logger.info(f"Parsing uploaded file: {file.filename}")
        
        # Read file bytes
        file_bytes = await file.read()
        
        # Determine document type from filename
        filename_lower = file.filename.lower()
        doc_type = "unknown"
        
        if filename_lower.endswith('.pdf'):
            doc_type = "pdf"
            content = extract_pdf_content(file_bytes)
        elif filename_lower.endswith('.docx') or filename_lower.endswith('.doc'):
            doc_type = "docx"
            content = extract_docx_content(file_bytes)
        elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
            doc_type = "xlsx"
            content = extract_xlsx_content(file_bytes)
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file type. Supported formats: PDF, DOCX, XLSX"
            )
        
        if not content.strip():
            raise HTTPException(
                status_code=400,
                detail="Document appears to be empty or content could not be extracted"
            )
        
        # Truncate content if too large
        MAX_CONTENT_LENGTH = 15000
        total_length = len(content)
        is_truncated = False
        
        if total_length > MAX_CONTENT_LENGTH:
            is_truncated = True
            first_part_length = int(MAX_CONTENT_LENGTH * 0.7)
            last_part_length = int(MAX_CONTENT_LENGTH * 0.1)
            
            content = (
                content[:first_part_length] + 
                f"\n\n... [Content truncated - showing first {first_part_length} and last {last_part_length} characters of {total_length} total] ...\n\n" +
                content[-last_part_length:]
            )
            logger.info(f"Truncated uploaded file content from {total_length} to {len(content)} characters")
        
        return ParseDocumentResponse(
            success=True,
            content=content,
            title=file.filename,
            doc_type=doc_type,
            is_truncated=is_truncated,
            total_length=total_length
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error parsing uploaded document: {e}")
        raise HTTPException(status_code=500, detail=f"Error parsing document: {str(e)}")
